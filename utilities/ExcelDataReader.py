import openpyxl


def get_test_data_as_dict(path):
    book = openpyxl.load_workbook(path)
    sheet = book.active
    registration_data = []

    for current_row in range(2, sheet.max_row + 1):
        current_dict = {}
        for current_column in range(1, sheet.max_column + 1):
            key = sheet.cell(row=1, column=current_column).value
            value = sheet.cell(row=current_row, column=current_column).value
            current_dict[key] = value
        registration_data.append(current_dict)
    return registration_data
